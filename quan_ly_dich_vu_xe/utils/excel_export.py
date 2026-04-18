import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

class ExcelExporter:
    @staticmethod
    def export_khach_hang(data, filename="khach_hang.xlsx"):
        """Xuất danh sách khách hàng ra Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sách khách hàng"
        
        # Style cho header
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4a9eff", end_color="4a9eff", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        headers = ["ID", "Mã KH", "Họ tên", "Số điện thoại", "Email", "Địa chỉ", "Ngày tạo"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Style cho dữ liệu
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row, kh in enumerate(data, 2):
            ws.cell(row=row, column=1, value=kh['id']).border = thin_border
            ws.cell(row=row, column=2, value=kh['ma_kh']).border = thin_border
            ws.cell(row=row, column=3, value=kh['ho_ten']).border = thin_border
            ws.cell(row=row, column=4, value=kh['so_dien_thoai']).border = thin_border
            ws.cell(row=row, column=5, value=kh['email'] or "").border = thin_border
            ws.cell(row=row, column=6, value=kh['dia_chi'] or "").border = thin_border
            
            # Xử lý ngày tháng
            ngay_tao = kh['ngay_tao']
            if isinstance(ngay_tao, datetime):
                date_str = ngay_tao.strftime("%d/%m/%Y")
            else:
                date_str = str(ngay_tao) if ngay_tao else ""
            ws.cell(row=row, column=7, value=date_str).border = thin_border
        
        # Auto-fit columns
        for col in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        wb.save(filename)
        return filename
    
    @staticmethod
    def export_xe(data, filename="xe.xlsx"):
        """Xuất danh sách xe ra Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sách xe"
        
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4a9eff", end_color="4a9eff", fill_type="solid")
        
        headers = ["ID", "Biển số", "Hiệu xe", "Model", "Màu sắc", "Năm SX", "Chủ xe"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row, x in enumerate(data, 2):
            ws.cell(row=row, column=1, value=x.get('id', ''))
            ws.cell(row=row, column=2, value=x.get('bien_so', ''))
            ws.cell(row=row, column=3, value=x.get('hieu_xe', ''))
            ws.cell(row=row, column=4, value=x.get('model', ''))
            ws.cell(row=row, column=5, value=x.get('mau_sac', ''))
            ws.cell(row=row, column=6, value=x.get('nam_sx', ''))
            ws.cell(row=row, column=7, value=x.get('ten_chu_xe', ''))
        
        for col in range(1, 8):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        wb.save(filename)
        return filename